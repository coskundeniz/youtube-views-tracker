from openpyxl import load_workbook

from exceptions import EmptyUrlListError
from utils import logger
from urlreader.urlreader import UrlReader


class ExcelReader(UrlReader):
    """Url reader for Excel files

    :type filename: str
    :param filename: Name of the file containing video urls
    :type url_column: int
    :param url_column: Url column index
    """

    def __init__(self, filename: str, url_column: int) -> None:

        self._filename = filename
        self._url_column = url_column

    def read_urls(self):
        """Read urls from file

        :rtype: list
        :returns: List of video urls
        """

        self.check_file_exists(self._filename)

        workbook = load_workbook(self._filename)
        sheet = workbook.active

        total_rows = len(list(sheet.rows))

        logger.info(f"Reading video urls from {self._filename}...")
        logger.debug(f"Number of rows: {total_rows}")

        urls = []

        for row_index in range(total_rows):
            cell = sheet.cell(row=row_index + 1, column=self._url_column + 1)
            urls.append(cell.value)

        if not all(urls):
            raise EmptyUrlListError("Video url list is empty! Please provide urls.")

        urls = self.check_url_count(urls)

        return urls
