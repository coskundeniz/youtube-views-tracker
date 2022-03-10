import os
import pytest

from openpyxl import load_workbook

from reporter.excel_reporter import ExcelReporter
from video import YoutubeVideo


@pytest.fixture
def reporter():

    output_file = "tests/reporter/results.xlsx"
    reporter = ExcelReporter(output_file)
    yield reporter

    if os.path.exists(output_file):
        os.remove(output_file)


@pytest.fixture
def videos():

    return [
        YoutubeVideo(
            url="https://www.youtube.com/watch?v=gxUq5Kt83V4",
            title="Darbuka solo before class",
            views=185,
        ),
        YoutubeVideo(
            url="https://www.youtube.com/watch?v=3T4uDDfR43Y",
            title="Thursday Memories",
            views=208,
        ),
        YoutubeVideo(
            url="https://www.youtube.com/watch?v=G0f9ms4tPjI",
            title="Dehollo - I",
            views=41,
        ),
    ]


def test_create_excel_report(reporter, videos):

    reporter.update_views(videos)

    workbook = load_workbook("tests/reporter/results.xlsx")
    sheet = workbook.active

    total_rows = len(list(sheet.rows))

    assert total_rows == 3
    assert sheet.cell(row=1, column=2).value == videos[0].title
    assert sheet.cell(row=2, column=3).value == videos[1].url
    assert sheet.cell(row=3, column=1).value == videos[2].views


def test_create_excel_report_with_existing_report(reporter, videos):

    reporter.update_views(videos)
    reporter.update_views(videos)

    workbook = load_workbook("tests/reporter/results.xlsx")
    sheet = workbook.active

    total_rows = len(list(sheet.rows))

    assert total_rows == 3
    assert sheet.cell(row=1, column=2).value == videos[0].title
    assert sheet.cell(row=2, column=3).value == videos[1].url
    assert sheet.cell(row=3, column=1).value == videos[2].views
