import os

import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

from utils import logger
from reporter.reporter import Reporter


class ExcelReporter(Reporter):
    """Excel report generator

    :type filename: str
    :param filename: Name of the output file
    """

    def __init__(self, filename: str) -> None:

        self._filename = filename

    def update_views(self, videos: list["YoutubeVideo"]) -> None:
        """Update view counts for videos on the output file

        :type videos: list
        :param videos: List of videos
        """

        logger.info(f"Updating view counts on {self._filename}...")

        workbook = self._create_workbook()

        sheet = workbook.active

        for index, video in enumerate(videos, start=1):

            views_cell = sheet.cell(row=index, column=1)
            title_cell = sheet.cell(row=index, column=2)
            url_cell = sheet.cell(row=index, column=3)

            views_cell.value = video.views
            title_cell.value = video.title
            url_cell.value = video.url

        titles = [video.title for video in videos]
        urls = [video.url for video in videos]
        self._adjust_column_width(sheet, "B", titles)
        self._adjust_column_width(sheet, "C", urls)

        self._prepare_chart_for_top_ten_videos(workbook)

        workbook.save(self._filename)

    def _prepare_chart_for_top_ten_videos(self, workbook: openpyxl.Workbook) -> None:
        """Prepare a bar chart on the second sheet for most watched 10 videos

        :type workbook: openpyxl.Workbook
        :param workbook: Workbook instance
        """

        chart_sheet_title = "Top 10 Videos Chart"
        if len(workbook.sheetnames) >= 2:
            chart_sheet = workbook[chart_sheet_title]
        else:
            chart_sheet = workbook.create_sheet(chart_sheet_title)

        chart = BarChart()
        chart.type = "col"
        chart.style = 4  # red
        chart.height = 18
        chart.width = 36
        chart.legend = None
        chart.title = "Most Watched 10 Videos"
        chart.x_axis.title = "Video Title"
        chart.y_axis.title = "View Count"

        data_sheet = self._get_sorted_worksheet(workbook)
        max_chart_items = 10
        total_rows = len(list(data_sheet.rows))
        max_rows = total_rows if total_rows <= max_chart_items else max_chart_items
        data = Reference(data_sheet, min_col=1, min_row=1, max_row=max_rows, max_col=2)
        categories = Reference(data_sheet, min_col=2, min_row=1, max_row=max_rows, max_col=2)
        chart.add_data(data)
        chart.set_categories(categories)
        chart_sheet.add_chart(chart, "B3")

    def _create_workbook(self) -> openpyxl.Workbook:
        """Create workbook

        If file exists, workbook is loaded by using it.

        :rtype: Workbook
        :returns: Workbook instance
        """

        if os.path.exists(self._filename):
            workbook = load_workbook(self._filename)
        else:
            workbook = openpyxl.Workbook()

        return workbook

    def _adjust_column_width(self, sheet: "Worksheet", column: str, contents: list[str]) -> None:
        """Adjust the width of given column according to maximum length of content

        :type sheet: Worksheet
        :param sheet: Active worksheet object
        :type column: str
        :param column: Column name
        :type contents: list
        :param contents: Column contents
        """

        max_content_length = max([len(content) for content in contents])
        sheet.column_dimensions[column].width = max_content_length + 2

    def _get_sorted_worksheet(self, workbook: openpyxl.Workbook) -> "Worksheet":
        """Get a worksheet sorted by view counts

        :type workbook: openpyxl.Workbook
        :param workbook: Workbook instance
        :rtype: Worksheet
        :returns: Sorted worksheet
        """

        current_sheet = workbook.active

        rows = []
        for row in current_sheet.iter_rows(values_only=True):
            rows.append(row)

        rows.sort(key=lambda x: x[0], reverse=True)

        sorted_data_sheet_title = "Sorted Data"

        # if sheet exists, remove before creating
        if sorted_data_sheet_title in workbook.sheetnames:
            del workbook[sorted_data_sheet_title]

        sorted_worksheet = workbook.create_sheet(sorted_data_sheet_title)

        for row in rows:
            sorted_worksheet.append(row)

        return sorted_worksheet
